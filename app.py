import streamlit as st
import pandas as pd
import openpyxl
import io

# 페이지 설정
st.set_page_config(page_title="정산 시스템", layout="wide")

st.title("📊 자동 비례 배부 정산 시스템")

# 1. 파일 업로드 및 데이터 불러오기
uploaded_file = st.file_uploader("엑셀 파일을 업로드하세요", type=["xlsx"])

if uploaded_file:
    df = pd.read_excel(uploaded_file)
    
    # 컬럼 설정 (엑셀 파일의 실제 컬럼명과 일치해야 합니다)
    dong_col = "동"
    ho_col = "호수"
    usage_col = "최초사용량"
    
    # 세션 상태 초기화
    if 'edited_df' not in st.session_state:
        st.session_state.edited_df = df.copy()
        st.session_state.edited_df['__fixed_row_id'] = range(len(st.session_state.edited_df))

    # [수정] 용어 변경: 전체 계량 종량 -> 월 전체 총량
    total_weight = st.number_input("월 전체 총량(kg)을 입력하세요:", min_value=0.0, step=0.1)

    # 3. 데이터 교정
    bad_rows = st.session_state.edited_df[st.session_state.edited_df[usage_col] > 999]
    
    if len(bad_rows) > 0:
        st.warning(f"🚨 검증 경고 ({len(bad_rows)}건 발견)")
        edited_bad_df = st.data_editor(bad_rows[[dong_col, ho_col, usage_col, '__fixed_row_id']], key="bad_df_editor")
        
        if st.button("💾 수정사항 저장 및 데이터 동기화"):
            for _, row in edited_bad_df.iterrows():
                target_id = int(row['__fixed_row_id'])
                st.session_state.edited_df.loc[st.session_state.edited_df['__fixed_row_id'] == target_id, usage_col] = row[usage_col]
            st.rerun()

    # 4. 정산 로직
    if st.button("정산 실행 및 엑셀 파일 생성"):
        df = st.session_state.edited_df
        
        # [핵심] 정밀 배부 로직
        total_initial_usage = df[usage_col].sum()
        # [수정] 변수명 및 로직 의미 명확화
        monthly_total_weight = total_weight 
        total_to_distribute = monthly_total_weight - total_initial_usage
        
        if total_initial_usage > 0:
            df['배부량'] = (df[usage_col] / total_initial_usage) * total_to_distribute
        else:
            df['배부량'] = monthly_total_weight / len(df)
            
        df['최종량'] = df[usage_col] + df['배부량']
        
        # [미세 오차 보정]
        error_adjustment = monthly_total_weight - df['최종량'].sum()
        df.iloc[0, df.columns.get_loc('최종량')] += error_adjustment
        
        # 5. 엑셀 생성
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "정산 결과"
        
        headers = ["동", "호수", "최초 사용량(kg)", "배부량(kg)", "최종 합계(kg)"]
        for i, h in enumerate(headers, 1):
            ws.cell(row=1, column=i, value=h)
        
        for i, row in df.iterrows():
            ws.cell(row=i+2, column=1, value=str(row[dong_col]))
            ws.cell(row=i+2, column=2, value=str(row[ho_col]))
            ws.cell(row=i+2, column=3, value=float(row[usage_col]))
            ws.cell(row=i+2, column=4, value=float(row['배부량']))
            ws.cell(row=i+2, column=5, value=float(row['최종량']))
            
        excel_data = io.BytesIO()
        wb.save(excel_data)
        excel_data.seek(0)
        
        st.success("✅ 정산 완료!")
        st.download_button("📁 정산 결과 다운로드", excel_data, "final_result.xlsx")
